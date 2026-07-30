#!/usr/bin/env python3
"""Build a deterministic XLSX mirror from the canonical CSV."""

import csv
import os
import tempfile
import zipfile
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

ROOT = Path(__file__).resolve().parents[1]
CSV_PATH = ROOT / "data" / "competitive-research-tracker.csv"
XLSX_PATH = ROOT / "data" / "competitive-research-tracker.xlsx"
FIXED_TIMESTAMP = (2026, 7, 30, 0, 0, 0)
ZIP_TIMESTAMP = (1980, 1, 1, 0, 0, 0)


def read_csv():
    with CSV_PATH.open(encoding="utf-8", newline="") as handle:
        reader = csv.reader(handle)
        rows = list(reader)
    if not rows:
        raise SystemExit("Canonical CSV is empty")
    width = len(rows[0])
    if any(len(row) != width for row in rows):
        raise SystemExit("Canonical CSV contains rows with inconsistent column counts")
    return rows


def write_workbook(rows, target):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Competitive Tracker"
    for row in rows:
        sheet.append(row)
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.sheet_view.showGridLines = True
    fixed = datetime(*FIXED_TIMESTAMP)
    workbook.properties.creator = "BizMatch Phase 0 canonical build"
    workbook.properties.lastModifiedBy = "BizMatch Phase 0 canonical build"
    workbook.properties.created = fixed
    workbook.properties.modified = fixed
    workbook.properties.title = "BizMatch Competitive Research Tracker"
    workbook.properties.description = (
        "Generated from data/competitive-research-tracker.csv; do not edit manually."
    )
    workbook.save(target)


def deterministic_repack(source, target):
    fd, repacked_name = tempfile.mkstemp(
        dir=target.parent, prefix=".competitive-tracker-", suffix=".xlsx"
    )
    os.close(fd)
    repacked = Path(repacked_name)
    try:
        with zipfile.ZipFile(source, "r") as src, zipfile.ZipFile(
            repacked, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=9
        ) as dst:
            for name in sorted(src.namelist()):
                info = zipfile.ZipInfo(name, ZIP_TIMESTAMP)
                info.compress_type = zipfile.ZIP_DEFLATED
                info.create_system = 3
                info.external_attr = 0o600 << 16
                dst.writestr(info, src.read(name))
        repacked.replace(target)
    finally:
        repacked.unlink(missing_ok=True)


def main():
    rows = read_csv()
    fd, raw_name = tempfile.mkstemp(
        dir=XLSX_PATH.parent, prefix=".competitive-tracker-raw-", suffix=".xlsx"
    )
    os.close(fd)
    raw = Path(raw_name)
    try:
        write_workbook(rows, raw)
        deterministic_repack(raw, XLSX_PATH)
    finally:
        raw.unlink(missing_ok=True)
    print(
        f"Generated {XLSX_PATH.relative_to(ROOT)} from {CSV_PATH.relative_to(ROOT)}: "
        f"{len(rows) - 1} records x {len(rows[0])} columns"
    )


if __name__ == "__main__":
    main()
