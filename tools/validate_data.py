#!/usr/bin/env python3
"""Phase 0 integrity checks for canonical data and generated artifacts."""

import argparse
import ast
import csv
import hashlib
import html
import json
import re
import subprocess
import sys
import time
import zipfile
from pathlib import Path
from urllib.parse import unquote, urlparse

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
CSV_PATH = ROOT / "data" / "competitive-research-tracker.csv"
XLSX_PATH = ROOT / "data" / "competitive-research-tracker.xlsx"
SUBSTITUTES_PATH = ROOT / "data" / "substitutes-research.csv"
SUBSTITUTE_EVIDENCE_PATH = ROOT / "data" / "substitute-evidence.csv"
SUBSTITUTE_WORKFLOWS_PATH = ROOT / "data" / "substitute-workflows.csv"
MARKET_RESEARCH_PATH = ROOT / "data" / "market-research.json"
MARKET_RESEARCH_PAGE = ROOT / "market-research" / "index.html"
FINDINGS_PATH = ROOT / "data" / "findings-and-implications.json"
FINDINGS_PAGE = ROOT / "findings-conclusions" / "index.html"
FINDINGS_REPORT = ROOT / "FINDINGS_AND_STRATEGIC_IMPLICATIONS.md"
PRESENTATION_PAGE = ROOT / "presentation" / "index.html"
SUBSTITUTE_REPORT_SCRIPT = ROOT / "tools" / "build_substitute_reports.py"
FINDINGS_REPORT_SCRIPT = ROOT / "tools" / "build_findings_report.py"
GENERATOR = ROOT / "tools" / "generate_site.py"
RECONCILIATION_SCRIPT = ROOT / "tools" / "apply_reconciliation.py"
EXPECTED_ROWS = 36
EXPECTED_COLUMNS = 65
EXPECTED_TRACKER_SHA256 = "cb6b808ab11198d2896a383e815c2b3c728000adfdeb785aafd4ec2882069a51"
MARKET_RESEARCH_TOP_LEVEL_KEYS = {
    "meta",
    "executive_conclusion",
    "research_inventory",
    "secondary_research_assessment",
    "interview_sample",
    "methodology_problems",
    "interview_findings",
    "hypothesis_register",
    "evidence_use_guidance",
    "evidence_register",
    "next_research_phase",
    "decision_gate_framework",
    "decision_gate",
    "sources",
}
MARKET_RESEARCH_HYPOTHESIS_STATUSES = {
    "Partially supported",
    "Partially supported by self-report",
    "Not supported",
    "Not tested",
    "Unresolved",
    "Unresolved for the target ICP",
    "Invalid evidence",
}
SUBSTITUTE_JOB_IDS = {
    "JOB-COFOUNDER",
    "JOB-FOUNDER-INVESTOR",
    "JOB-INVESTOR-SOURCING",
    "JOB-TRUSTED-PROGRESSION",
}
SUBSTITUTE_STAGE_IDS = {
    "need_definition",
    "discovery",
    "initial_screen",
    "fit_check",
    "contact",
    "response_followup",
    "trust_building",
    "verification",
    "disclosure",
    "nda",
    "meeting",
    "diligence",
    "decision_followup",
}
SUBSTITUTE_CLASSIFICATIONS = {
    "Direct Competitor",
    "Adjacent Competitor",
    "Workflow Substitute",
    "Service Substitute",
    "Community/Network Substitute",
    "Manual Process",
    "Do Nothing",
    "Infrastructure Tool",
    "Unclear",
}
SUBSTITUTE_STRENGTHS = {
    "Strong Substitute",
    "Partial Substitute",
    "Weak Substitute",
    "Complementary Tool",
    "Insufficient Evidence",
}
SUBSTITUTE_SOURCE_TYPES = {
    "Independent Behavioral Evidence",
    "Independent User Report",
    "Independent Market Evidence",
    "Company Documentation",
    "Company Claim",
    "Community Discussion",
    "Anecdotal Evidence",
    "Inference",
    "Unverified",
}
FINDINGS_CONFIDENCE = {"High", "Medium", "Low", "Insufficient Evidence"}
SUBSTITUTE_HEADERS = (
    "substitute_id",
    "name",
    "category",
    "classification",
    "target_persona",
    "job_to_be_done",
    "workflow_stages_covered",
    "current_behavior",
    "why_users_choose_it",
    "advantages",
    "limitations",
    "switching_cost",
    "trust_mechanism",
    "payment_model",
    "substitute_strength",
    "evidence_summary",
    "evidence_ids",
    "source_url",
    "source_title",
    "source_type",
    "source_date",
    "last_verified",
    "confidence",
    "research_status",
    "existing_competitor_slug",
    "notes",
)
SUBSTITUTE_EVIDENCE_HEADERS = (
    "evidence_id",
    "substitute_id",
    "job_id",
    "claim_type",
    "evidence_dimension",
    "claim",
    "source_url",
    "source_title",
    "source_date",
    "last_verified",
    "source_type",
    "supporting_excerpt_or_summary",
    "confidence",
    "limitation",
)
SUBSTITUTE_WORKFLOW_HEADERS = (
    "workflow_id",
    "job_id",
    "stage_id",
    "stage_order",
    "current_action",
    "tools_channels",
    "people_involved",
    "manual_work",
    "time_or_friction",
    "trust_requirement",
    "current_advantage",
    "current_limitation",
    "substitute_ids",
    "evidence_ids",
    "evidence_status",
    "confidence",
)
LEGACY_FIELDS = {
    "product_overlap_score",
    "feature_maturity_score",
    "market_traction_score",
    "funding_strength_score",
    "ai_depth_score",
    "nda_security_strength_score",
    "network_moat_score",
    "direct_threat_score",
}
RECONCILED_FIELDS = {
    ("Comatch", "total_funding"),
    ("FounderCloud", "hq_country"),
    ("FounderCloud", "notes"),
    ("Gust", "users_traction"),
    ("OpenVC", "users_traction"),
    ("Visible.vc", "hq_country"),
    ("Visible.vc", "total_funding"),
    ("Peachscore", "total_funding"),
    ("Peachscore", "funding_rounds"),
}
PROFILE_FIELDS = {
    "Comatch": ("total_funding",),
    "FounderCloud": ("hq_country", "notes"),
    "Gust": ("users_traction",),
    "OpenVC": ("users_traction",),
    "Visible.vc": ("hq_country", "total_funding"),
    "Peachscore": ("total_funding", "funding_rounds"),
}


class Validation:
    def __init__(self):
        self.failures = []
        self.passes = []

    def check(self, condition, message):
        if condition:
            self.passes.append(message)
        else:
            self.failures.append(message)

    def require(self):
        if self.failures:
            for failure in self.failures:
                print(f"FAIL: {failure}", file=sys.stderr)
            raise SystemExit(1)


def slug(value):
    return re.sub(r"-+", "-", re.sub(r"[^a-z0-9]+", "-", value.lower())).strip("-")


def read_canonical():
    with CSV_PATH.open(encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle)
        headers = reader.fieldnames or []
        rows = list(reader)
    return headers, rows


def read_csv_table(path):
    with path.open(encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle)
        return tuple(reader.fieldnames or ()), list(reader)


def split_values(value):
    return [item.strip() for item in (value or "").split("|") if item.strip()]


def normalized_name(value):
    return re.sub(r"[^a-z0-9]+", "", (value or "").lower())


def extract_urls(text):
    return [match.rstrip(").,") for match in re.findall(r"https?://[^\s;]+", text or "")]


def normalized(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def check_canonical(validation, headers, rows):
    validation.check(CSV_PATH.exists(), "canonical CSV exists")
    validation.check(
        hashlib.sha256(CSV_PATH.read_bytes()).hexdigest() == EXPECTED_TRACKER_SHA256,
        "Phase 0 canonical competitor tracker remains byte-for-byte unchanged",
    )
    active_csvs = sorted(path.name for path in (ROOT / "data").glob("*.csv"))
    expected_csvs = sorted(
        (
            CSV_PATH.name,
            SUBSTITUTES_PATH.name,
            SUBSTITUTE_EVIDENCE_PATH.name,
            SUBSTITUTE_WORKFLOWS_PATH.name,
        )
    )
    validation.check(
        active_csvs == expected_csvs,
        "data/ contains only the separated canonical competitor and substitute CSV layers",
    )
    validation.check(len(rows) == EXPECTED_ROWS, "canonical CSV contains 36 companies")
    validation.check(len(headers) == EXPECTED_COLUMNS, "canonical CSV contains 65 columns")
    validation.check(len(headers) == len(set(headers)), "canonical headers are unique")
    companies = [row.get("company", "").strip() for row in rows]
    validation.check(all(companies), "all company identifiers are non-empty")
    validation.check(len(companies) == len(set(companies)), "company identifiers are unique")
    slugs = [slug(company) for company in companies]
    validation.check(len(slugs) == len(set(slugs)), "generated company slugs are unique")

    malformed = []
    for row in rows:
        for field in ("url", "primary_sources", "secondary_sources"):
            value = row.get(field, "")
            found = extract_urls(value)
            if field == "url" and value and not found:
                malformed.append(f"{row['company']}.{field}: {value}")
            for source_url in found:
                parsed = urlparse(source_url)
                if parsed.scheme not in {"http", "https"} or not parsed.netloc:
                    malformed.append(f"{row['company']}.{field}: {source_url}")
    validation.check(not malformed, "all stored source links are structurally valid")

    by_company = {row["company"]: row for row in rows}
    missing_audit = []
    for company, field in RECONCILED_FIELDS:
        row = by_company[company]
        has_source = bool(extract_urls(row["primary_sources"] + " " + row["secondary_sources"]))
        explicitly_unverified = any(
            marker in row[field] for marker in ("Unresolved", "Insufficient Evidence", "Company Claim")
        )
        if not (has_source or explicitly_unverified):
            missing_audit.append(f"{company}.{field}")
    validation.check(
        not missing_audit,
        "every reconciled field has a source or explicit unresolved marker",
    )
    legacy_narrative = []
    for row in rows:
        for field, value in row.items():
            if field in LEGACY_FIELDS:
                continue
            for legacy_field in LEGACY_FIELDS:
                if legacy_field in (value or ""):
                    legacy_narrative.append(f"{row['company']}.{field}:{legacy_field}")
    validation.check(
        not legacy_narrative,
        "legacy score names do not appear in active research narratives",
    )
    return by_company


def check_substitutes(validation, competitor_rows, artifacts_required):
    tables = (
        (SUBSTITUTES_PATH, SUBSTITUTE_HEADERS),
        (SUBSTITUTE_EVIDENCE_PATH, SUBSTITUTE_EVIDENCE_HEADERS),
        (SUBSTITUTE_WORKFLOWS_PATH, SUBSTITUTE_WORKFLOW_HEADERS),
    )
    loaded = {}
    for path, expected_headers in tables:
        validation.check(path.exists(), f"{path.name} exists")
        headers, rows = read_csv_table(path)
        loaded[path] = (headers, rows)
        validation.check(
            headers == expected_headers,
            f"{path.name} uses the documented canonical schema and column order",
        )
        validation.check(
            len(headers) == len(set(headers)),
            f"{path.name} has unique headers",
        )

    substitute_headers, substitutes = loaded[SUBSTITUTES_PATH]
    evidence_headers, evidence = loaded[SUBSTITUTE_EVIDENCE_PATH]
    workflow_headers, workflows = loaded[SUBSTITUTE_WORKFLOWS_PATH]
    del substitute_headers, evidence_headers, workflow_headers

    substitute_ids = [row["substitute_id"].strip() for row in substitutes]
    names = [row["name"].strip() for row in substitutes]
    normalized_names = [normalized_name(name) for name in names]
    validation.check(
        all(substitute_ids) and len(substitute_ids) == len(set(substitute_ids)),
        "substitute identifiers are non-empty and unique",
    )
    validation.check(
        all(names) and len(normalized_names) == len(set(normalized_names)),
        "substitute names remain unique after punctuation and case normalization",
    )
    validation.check(
        all(row["classification"] in SUBSTITUTE_CLASSIFICATIONS for row in substitutes),
        "every substitute has an allowed classification",
    )
    validation.check(
        all(
            split_values(row["job_to_be_done"])
            and set(split_values(row["job_to_be_done"])) <= SUBSTITUTE_JOB_IDS
            for row in substitutes
        ),
        "every substitute links to one or more allowed Jobs",
    )
    validation.check(
        all(row["substitute_strength"] in SUBSTITUTE_STRENGTHS for row in substitutes),
        "substitute strength is qualitative and uses only the documented categories",
    )
    validation.check(
        not any("score" in header.lower() for header in SUBSTITUTE_HEADERS),
        "substitute entity schema contains no numeric score field",
    )

    competitor_slugs = {slug(row["company"]): row["company"] for row in competitor_rows}
    linked_errors = []
    for row in substitutes:
        link_slug = row["existing_competitor_slug"].strip()
        name_slug = slug(row["name"].split(" such as ")[0])
        if link_slug and link_slug not in competitor_slugs:
            linked_errors.append(f"{row['substitute_id']}: unknown {link_slug}")
        exact_competitor = next(
            (
                company_slug
                for company_slug, company in competitor_slugs.items()
                if normalized_name(company) in normalized_name(row["name"])
            ),
            None,
        )
        if exact_competitor and not link_slug:
            linked_errors.append(
                f"{row['substitute_id']}: existing competitor {exact_competitor} is not linked"
            )
        if link_slug and link_slug != name_slug and normalized_name(
            competitor_slugs[link_slug]
        ) not in normalized_name(row["name"] + " " + row["notes"]) and link_slug != "yc-co-founder-matching":
            linked_errors.append(f"{row['substitute_id']}: ambiguous link {link_slug}")
    validation.check(
        not linked_errors,
        "existing competitor entities are linked by canonical slug and not silently duplicated",
    )

    malformed = []
    source_gaps = []
    for row in substitutes:
        source_type = row["source_type"].strip()
        source_url = row["source_url"].strip()
        if source_type not in SUBSTITUTE_SOURCE_TYPES:
            source_gaps.append(f"{row['substitute_id']}: invalid source type {source_type}")
        if source_type == "Unverified":
            if row["research_status"] != "Unverified":
                source_gaps.append(f"{row['substitute_id']}: unverified source not reflected in status")
        elif not source_url:
            source_gaps.append(f"{row['substitute_id']}: sourced record has no URL")
        if source_url:
            parsed = urlparse(source_url)
            if parsed.scheme not in {"http", "https"} or not parsed.netloc:
                malformed.append(f"{row['substitute_id']}: {source_url}")
    validation.check(
        not source_gaps,
        "every substitute has a source or is explicitly Unverified",
    )
    validation.check(
        not malformed,
        "all substitute representative source URLs are structurally valid",
    )

    evidence_ids = [row["evidence_id"].strip() for row in evidence]
    validation.check(
        all(evidence_ids) and len(evidence_ids) == len(set(evidence_ids)),
        "substitute evidence identifiers are non-empty and unique",
    )
    evidence_id_set = set(evidence_ids)
    evidence_source_errors = []
    evidence_link_errors = []
    for row in evidence:
        source_type = row["source_type"].strip()
        source_url = row["source_url"].strip()
        if source_type not in SUBSTITUTE_SOURCE_TYPES:
            evidence_source_errors.append(
                f"{row['evidence_id']}: invalid source type {source_type}"
            )
        if not row["claim"].strip():
            evidence_source_errors.append(f"{row['evidence_id']}: empty claim")
        if source_type == "Unverified":
            if source_url:
                evidence_source_errors.append(
                    f"{row['evidence_id']}: Unverified row unexpectedly has a URL"
                )
        elif not source_url:
            evidence_source_errors.append(f"{row['evidence_id']}: sourced claim has no URL")
        if source_url:
            parsed = urlparse(source_url)
            if parsed.scheme not in {"http", "https"} or not parsed.netloc:
                evidence_source_errors.append(
                    f"{row['evidence_id']}: malformed URL {source_url}"
                )
        if source_type == "Inference" and row["evidence_dimension"] != "Inference":
            evidence_source_errors.append(
                f"{row['evidence_id']}: Inference is not explicitly labeled"
            )
        unknown_substitutes = set(split_values(row["substitute_id"])) - set(substitute_ids)
        unknown_jobs = set(split_values(row["job_id"])) - SUBSTITUTE_JOB_IDS
        if unknown_substitutes or unknown_jobs:
            evidence_link_errors.append(
                f"{row['evidence_id']}: substitutes={sorted(unknown_substitutes)} "
                f"jobs={sorted(unknown_jobs)}"
            )
    validation.check(
        not evidence_source_errors,
        "every factual evidence claim has a valid source and explicit evidence type",
    )
    validation.check(
        not evidence_link_errors,
        "evidence records link only to canonical substitutes and Jobs",
    )

    orphan_references = []
    for row in substitutes:
        unknown = set(split_values(row["evidence_ids"])) - evidence_id_set
        if unknown:
            orphan_references.append(
                f"{row['substitute_id']}: {sorted(unknown)}"
            )
    workflow_ids = [row["workflow_id"].strip() for row in workflows]
    validation.check(
        all(workflow_ids) and len(workflow_ids) == len(set(workflow_ids)),
        "workflow identifiers are non-empty and unique",
    )
    workflow_errors = []
    coverage = {job_id: set() for job_id in SUBSTITUTE_JOB_IDS}
    stage_coverage = {job_id: set() for job_id in SUBSTITUTE_JOB_IDS}
    for row in workflows:
        job_id = row["job_id"].strip()
        if job_id not in SUBSTITUTE_JOB_IDS:
            workflow_errors.append(f"{row['workflow_id']}: invalid Job {job_id}")
            continue
        try:
            stage_order = int(row["stage_order"])
        except ValueError:
            workflow_errors.append(f"{row['workflow_id']}: invalid stage order")
            continue
        coverage[job_id].add(stage_order)
        stage_coverage[job_id].add(row["stage_id"].strip())
        unknown_substitutes = set(split_values(row["substitute_ids"])) - set(substitute_ids)
        unknown_evidence = set(split_values(row["evidence_ids"])) - evidence_id_set
        if unknown_substitutes or unknown_evidence:
            workflow_errors.append(
                f"{row['workflow_id']}: substitutes={sorted(unknown_substitutes)} "
                f"evidence={sorted(unknown_evidence)}"
            )
        if not row["evidence_ids"].strip() and row["evidence_status"] != "Unverified":
            workflow_errors.append(
                f"{row['workflow_id']}: missing evidence is not marked Unverified"
            )
    validation.check(
        not workflow_errors,
        "workflow rows use canonical references and mark unsupported stages Unverified",
    )
    validation.check(
        all(orders == set(range(1, 14)) for orders in coverage.values()),
        "all four Jobs contain exactly the thirteen documented workflow stages",
    )
    validation.check(
        all(stages == SUBSTITUTE_STAGE_IDS for stages in stage_coverage.values()),
        "every Job uses each documented workflow-stage identifier exactly once",
    )
    validation.check(
        not orphan_references,
        "substitute entity evidence references resolve to the evidence register",
    )

    tracker_text = CSV_PATH.read_text(encoding="utf-8")
    validation.check(
        "substitute_id" not in tracker_text
        and not any(
            row["substitute_id"] in tracker_text
            for row in substitutes
        ),
        "substitute records are not injected into the 36-company tracker",
    )

    score_source = function_source(GENERATOR, "relationship_score")
    validation.check(
        not any(
            repr(row["name"]) in score_source or f'"{row["name"]}"' in score_source
            for row in substitutes
        ),
        "relationship scoring contains no substitute-name assignments",
    )
    renderer_source = function_source(GENERATOR, "substitutes_page")
    report_source = SUBSTITUTE_REPORT_SCRIPT.read_text(encoding="utf-8")
    validation.check(
        not any(
            row["name"] in renderer_source or row["name"] in report_source
            for row in substitutes
        ),
        "substitute outputs are data-driven and contain no hard-coded substitute list",
    )

    if not artifacts_required:
        return
    substitute_page = ROOT / "sites" / "full-report-site" / "alternative-workflows.html"
    substitute_data = ROOT / "sites" / "full-report-site" / "substitutes-data.js"
    generated_reports = (
        ROOT / "SUBSTITUTE_MATRIX.md",
        ROOT / "SUBSTITUTE_WORKFLOWS.md",
        ROOT / "SUBSTITUTE_EVIDENCE_REGISTER.md",
    )
    validation.check(
        substitute_page.exists() and substitute_data.exists(),
        "site includes the generated Alternative Workflows data and page",
    )
    validation.check(
        all(path.exists() for path in generated_reports),
        "generated substitute matrix, workflow maps, and evidence register exist",
    )
    if substitute_page.exists() and substitute_data.exists():
        page_text = substitute_page.read_text(encoding="utf-8")
        data_text = substitute_data.read_text(encoding="utf-8")
        validation.check(
            all(row["substitute_id"] in data_text for row in substitutes)
            and "data/substitutes-research.csv" in data_text,
            "Alternative Workflows data is generated from every canonical substitute record",
        )
        labeled_types = {
            row["source_type"]
            for row in evidence
            if row["source_type"] in {"Company Claim", "Inference", "Unverified"}
        }
        validation.check(
            all(source_type in page_text for source_type in labeled_types)
            and "not independent proof of effectiveness" in page_text,
            "Company Claim, Inference, and Unverified evidence remain visibly labeled",
        )
        validation.check(
            "relationship_score" not in data_text
            and "direct_threat_score" not in data_text,
            "substitute site data contains no numeric relationship or legacy threat score",
        )


def nested_keys(value):
    if isinstance(value, dict):
        for key, child in value.items():
            yield key
            yield from nested_keys(child)
    elif isinstance(value, list):
        for child in value:
            yield from nested_keys(child)


def check_market_research(validation, artifacts_required):
    validation.check(
        MARKET_RESEARCH_PATH.exists(),
        "separate canonical market-research content source exists",
    )
    if not MARKET_RESEARCH_PATH.exists():
        return
    try:
        research = json.loads(MARKET_RESEARCH_PATH.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, UnicodeDecodeError):
        validation.check(False, "market-research content is valid UTF-8 JSON")
        return
    validation.check(True, "market-research content is valid UTF-8 JSON")
    validation.check(
        set(research) == MARKET_RESEARCH_TOP_LEVEL_KEYS,
        "market-research content uses the documented top-level structure",
    )

    sample = research.get("interview_sample", {})
    stages = sample.get("stages", [])
    expected_stages = {
        "Idea": 3,
        "MVP": 3,
        "Product": 2,
        "Customers": 3,
        "Revenue": 3,
        "Scale": 1,
    }
    actual_stages = {
        row.get("stage"): row.get("interviews")
        for row in stages
        if isinstance(row, dict)
    }
    validation.check(
        actual_stages == expected_stages
        and sum(expected_stages.values()) == sample.get("total") == 15,
        "interview stage distribution preserves the supplied 15-participant counts",
    )
    validation.check(
        sample.get("active_partner_seekers") == 2
        and sample.get("later_stage_participants") == 7,
        "sample assessment preserves the supplied active-seeker and later-stage counts",
    )

    findings = research.get("interview_findings", {})
    primary_pain = findings.get("primary_pain", {})
    primary_denominator = primary_pain.get("denominator", {})
    validation.check(
        primary_pain.get("headline") == "Question-level denominator not documented"
        and primary_pain.get("label")
        == "No participant who answered the primary-pain question identified finding a co-founder as the single problem they would remove."
        and primary_denominator.get("asked") == "Not documented"
        and primary_denominator.get("answered") == "Not documented"
        and primary_denominator.get("usable") == "Requires recoding"
        and "0/15" not in json.dumps(primary_pain),
        "primary-pain finding does not manufacture a question-level denominator",
    )
    partner_origin = findings.get("partner_origin", {})
    channel_counts = {
        row.get("channel"): row.get("count")
        for row in partner_origin.get("channels", [])
        if isinstance(row, dict)
    }
    validation.check(
        partner_origin.get("usable_interviews") == 10
        and partner_origin.get("excluded_record_status") == "Invalid evidence"
        and sum(channel_counts.values()) == 10
        and channel_counts.get("Friends") == 3
        and channel_counts.get("Previous work") == 3
        and channel_counts.get("Studies, programs, or another existing relationship") == 4
        and channel_counts.get("Dedicated matching platform") == 0,
        "partner-origin findings preserve the supplied valid-record arithmetic",
    )
    validation.check(
        partner_origin.get("provisional_note")
        == "Provisional pending recoding: The original interviews did not consistently distinguish co-founders from other partner types."
        and partner_origin.get("conclusion")
        == "In the ten currently usable retrospective records, partners were found through existing relationships rather than dedicated matching platforms. This suggests that warm networks may be an important substitute, but co-founder-specific recoding and research with active seekers are still required."
        and partner_origin.get("denominator", {}).get("usable")
        == "10 retrospectively coded records",
        "partner-origin interpretation is provisional, bounded, and denominator-aware",
    )
    disclosure = findings.get("disclosure", {})
    validation.check(
        disclosure.get("would_not_disclose_openly") == 12
        and disclosure.get("reported_no_restriction") == 1
        and disclosure.get("not_asked_or_no_answer") == 2
        and disclosure.get("explicit_nda_mentions") == 1
        and (
            disclosure.get("would_not_disclose_openly", 0)
            + disclosure.get("reported_no_restriction", 0)
            + disclosure.get("not_asked_or_no_answer", 0)
            == 15
        ),
        "disclosure and NDA findings preserve the supplied interview counts",
    )
    validation.check(
        disclosure.get("conclusion")
        == "Selective disclosure was common in self-report. The research did not establish that it delayed introductions, prevented meetings, caused abandonment or created demand for an NDA workflow."
        and disclosure.get("denominator", {}).get("answered") == "13 coded responses"
        and disclosure.get("denominator", {}).get("missing_or_not_asked")
        == "2 records",
        "disclosure finding separates self-report from behavioral severity and NDA demand",
    )

    hypotheses = research.get("hypothesis_register", [])
    validation.check(
        len(hypotheses) == 14
        and all(
            row.get("status") in MARKET_RESEARCH_HYPOTHESIS_STATUSES
            for row in hypotheses
        ),
        "hypothesis register uses qualitative evidence statuses only",
    )
    urgency = next(
        (
            row
            for row in hypotheses
            if row.get("hypothesis") == "The problem is urgent and recurring"
        ),
        {},
    )
    validation.check(
        urgency.get("status") == "Unresolved for the target ICP"
        and urgency.get("detail")
        == "Co-founder search did not emerge as a primary pain in the current broad convenience sample, but the sample contained only two active partner seekers and cannot resolve urgency among Validation/MVP founders actively searching for a co-founder.",
        "urgency remains a negative observation but unresolved for the target ICP",
    )

    evidence_register = research.get("evidence_register", [])
    required_claim_ids = {f"MR-CLM-{number:03d}" for number in range(1, 8)}
    required_claim_fields = {
        "claim_id",
        "claim",
        "question_or_evidence_source",
        "population",
        "denominator",
        "coding_rule",
        "exclusions",
        "evidence_type",
        "research_date",
        "private_source_reference",
        "confidence_limitation",
    }
    claim_ids = [row.get("claim_id") for row in evidence_register]
    validation.check(
        len(evidence_register) == 7
        and set(claim_ids) == required_claim_ids
        and len(claim_ids) == len(set(claim_ids))
        and all(set(row) == required_claim_fields for row in evidence_register)
        and all(
            row.get("evidence_type") in {"Self-report", "Secondary source", "Inference"}
            for row in evidence_register
        ),
        "anonymized evidence register covers the seven required claims with stable fields",
    )
    validation.check(
        all(
            any(
                marker in " ".join(
                    str(row.get(field, ""))
                    for field in (
                        "denominator",
                        "coding_rule",
                        "exclusions",
                        "research_date",
                        "private_source_reference",
                        "confidence_limitation",
                    )
                )
                for marker in (
                    "Not documented",
                    "Requires recoding",
                    "Observed in current sample",
                    "Partially supported by self-report",
                    "Provisional pending recoding",
                )
            )
            for row in evidence_register
        ),
        "evidence-register uncertainty is expressed in plain language rather than scores",
    )

    next_phases = research.get("next_research_phase", [])
    validation.check(
        [phase.get("phase") for phase in next_phases]
        == ["Phase A", "Phase B", "Phase C", "Phase D", "Phase E"]
        and [phase.get("title") for phase in next_phases]
        == [
            "Recode the 15 existing interviews",
            "Focused founder discovery",
            "Separate investor research",
            "Behavioural concierge test",
            "Disclosure and NDA experiment",
        ],
        "Next Research Phase contains the specified five operational phases",
    )
    framework = research.get("decision_gate_framework", {})
    validation.check(
        framework.get("threshold_notice")
        == "Numerical decision thresholds must be approved before recruitment begins. They must not be selected after seeing the results."
        and len(framework.get("metrics_to_collect", [])) >= 5
        and len(framework.get("rows", [])) == 4
        and all(
            set(row) == {"observation", "evidence_required", "decision_permitted"}
            for row in framework.get("rows", [])
        )
        and any(
            "No broad two-sided marketplace build is justified."
            == row.get("decision_permitted")
            for row in framework.get("rows", [])
        ),
        "decision-gate framework requires pre-approved thresholds and permits only bounded next steps",
    )
    all_keys = set(nested_keys(research))
    validation.check(
        not any(
            re.search(r"(?:score|rating|percentage)$", key, flags=re.IGNORECASE)
            for key in all_keys
        ),
        "market research defines no numerical validation or decision score",
    )

    malformed_sources = []
    for source in research.get("sources", []):
        parsed = urlparse(source.get("url", ""))
        if parsed.scheme not in {"http", "https"} or not parsed.netloc:
            malformed_sources.append(source.get("url", ""))
    validation.check(
        len(research.get("sources", [])) == 4 and not malformed_sources,
        "market-research source register contains the four supplied structural URLs",
    )

    private_keys = {
        "participant_name",
        "participant_names",
        "email",
        "phone",
        "contact_details",
        "recording_url",
        "transcript",
        "transcript_url",
    }
    raw_content = MARKET_RESEARCH_PATH.read_text(encoding="utf-8")
    validation.check(
        not (all_keys & private_keys)
        and not re.search(r"\b[\w.+-]+@[\w.-]+\.[A-Za-z]{2,}\b", raw_content)
        and not re.search(r"\b05\d[- ]?\d{3}[- ]?\d{4}\b", raw_content),
        "public market-research content contains no participant PII fields or contact values",
    )
    validation.check(
        not re.search(
            r"(?:drive\.google\.com|docs\.google\.com|linkedin\.com/in/|"
            r"(?:recording|transcript)[_-]?url)",
            raw_content,
            flags=re.IGNORECASE,
        ),
        "public market-research source contains no private document, social-profile, recording, or transcript link",
    )

    if not artifacts_required:
        return
    validation.check(
        MARKET_RESEARCH_PAGE.exists(),
        "stable /market-research/ page is generated",
    )
    if not MARKET_RESEARCH_PAGE.exists():
        return
    page_text = MARKET_RESEARCH_PAGE.read_text(encoding="utf-8")
    required_sections = (
        "Existing research inventory",
        "Secondary research assessment",
        "Interview sample assessment",
        "Interview-methodology problems",
        "What the interviews actually show",
        "Hypothesis register",
        "Evidence-use guidance",
        "Evidence Register",
        "Next Research Phase",
        "Decision Gate Before Product Expansion",
        "Sources and interpretation limits",
    )
    validation.check(
        all(section in page_text for section in required_sections),
        "Market Research page contains every required analytical section",
    )
    required_boundaries = (
        "Exploratory research; not market validation",
        "Narrow the research",
        "Question-level denominator not documented",
        "Research scope",
        "Investor demand, project discovery, willingness to pay and the dynamics of a two-sided marketplace remain untested.",
        "Provisional pending recoding",
        "Unresolved for the target ICP",
        "Selective disclosure was common in self-report.",
        "Numerical decision thresholds must be approved before recruitment begins.",
        "Scenario boundary",
        "Hypothesis requiring validation",
        "Statuses are qualitative evidence labels, not numerical scores.",
        "No private interview artifact is linked",
    )
    validation.check(
        all(boundary in page_text for boundary in required_boundaries)
        and "does not yet justify Proceed, Pivot, or Stop" in page_text,
        "Market Research page distinguishes evidence, hypothesis, uncertainty, and decision",
    )
    required_section_ids = (
        "executive-conclusion",
        "research-inventory",
        "secondary-research",
        "interview-sample",
        "methodology-risks",
        "findings",
        "hypothesis-register",
        "evidence-register",
        "next-research",
        "decision-gate",
        "sources",
    )
    validation.check(
        'aria-label="On this page"' in page_text
        and all(
            f'id="{section_id}"' in page_text
            and f'href="#{section_id}"' in page_text
            for section_id in required_section_ids
        ),
        "Market Research page provides accessible navigation to stable section anchors",
    )
    validation.check(
        all(claim_id in page_text for claim_id in required_claim_ids)
        and page_text.count('class="research-phase"') == 5
        and "mobile-card-table" in page_text
        and 'data-label="Denominator"' in page_text,
        "evidence register, five-phase plan, and responsive table metadata are rendered",
    )
    validation.check(
        "0/15" not in page_text
        and "Not supported by the broad sample" not in page_text
        and "The current alternative is not merely LinkedIn or WhatsApp" not in page_text,
        "unverified denominator and overgeneralized market conclusions are absent",
    )
    footer = page_text.split('<footer class="footer market-footer">', 1)[-1]
    validation.check(
        '<footer class="footer market-footer">' in page_text
        and "Competitor tracker" not in footer
        and "Substitute entities" not in footer
        and "numeric threat ranking" not in footer,
        "Market Research footer excludes competitor and substitute research material",
    )
    validation.check(
        all(source["url"] in page_text for source in research.get("sources", [])),
        "all supplied external source links are rendered on the Market Research page",
    )
    validation.check(
        not re.search(r"\b[\w.+-]+@[\w.-]+\.[A-Za-z]{2,}\b", page_text)
        and not re.search(r"\b05\d[- ]?\d{3}[- ]?\d{4}\b", page_text)
        and "mailto:" not in page_text,
        "generated Market Research page exposes no participant contact information",
    )
    validation.check(
        "drive.google.com" not in page_text
        and "docs.google.com" not in page_text
        and "linkedin.com/in/" not in page_text
        and not re.search(r"https?://[^\"']*(?:recording|transcript)", page_text, re.I),
        "generated Market Research page exposes no private research or participant-profile links",
    )

    full_report_pages = list((ROOT / "sites" / "full-report-site").rglob("*.html"))
    validation.check(
        bool(full_report_pages)
        and all(
            'href="/market-research/"' in path.read_text(encoding="utf-8")
            and ">Market Research</a>" in path.read_text(encoding="utf-8")
            for path in full_report_pages
        )
        and 'aria-current="page">Market Research</a>' in page_text,
        "Market Research is a keyboard-accessible top-level tab across existing routes",
    )
    netlify_config = (ROOT / "netlify.toml").read_text(encoding="utf-8")
    validation.check(
        'from = "/market-research"' in netlify_config
        and 'to = "/market-research/index.html"' in netlify_config
        and "status = 200" in netlify_config,
        "Netlify direct loading and refresh are configured for /market-research",
    )


def walk_dicts(value):
    if isinstance(value, dict):
        yield value
        for child in value.values():
            yield from walk_dicts(child)
    elif isinstance(value, list):
        for child in value:
            yield from walk_dicts(child)


def check_findings(validation, competitor_rows, artifacts_required):
    validation.check(
        FINDINGS_PATH.exists(),
        "separate canonical active-findings source exists",
    )
    if not FINDINGS_PATH.exists():
        return
    try:
        findings = json.loads(FINDINGS_PATH.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, UnicodeDecodeError):
        validation.check(False, "active-findings content is valid UTF-8 JSON")
        return
    validation.check(True, "active-findings content is valid UTF-8 JSON")
    required_top_level = {
        "meta",
        "executive_conclusion",
        "jobs",
        "cross_market_findings",
        "competitive_pressures",
        "assumptions_strengthened",
        "assumptions_weakened",
        "pains",
        "conclusion_boundaries",
        "customer_discovery",
        "conditional_implications",
        "red_team",
        "methodology_limitations",
    }
    validation.check(
        set(findings) == required_top_level,
        "active findings use the documented structured conclusion schema",
    )

    _, evidence_rows = read_csv_table(SUBSTITUTE_EVIDENCE_PATH)
    _, substitute_rows = read_csv_table(SUBSTITUTES_PATH)
    evidence_ids = {row["evidence_id"] for row in evidence_rows}
    substitute_ids = {row["substitute_id"] for row in substitute_rows}
    competitor_slugs = {slug(row["company"]) for row in competitor_rows}
    market = json.loads(MARKET_RESEARCH_PATH.read_text(encoding="utf-8"))
    hypothesis_names = {
        row["hypothesis"] for row in market.get("hypothesis_register", [])
    }
    inventory_areas = {
        row["area"] for row in market.get("research_inventory", {}).get("rows", [])
    }

    reference_errors = []
    confidence_errors = []
    material_without_trace = []
    material_markers = {
        "headline",
        "finding",
        "pressure",
        "assumption",
        "pain",
        "evidence_observed",
        "possibility",
        "dominant_current_workflow",
    }
    for item in walk_dicts(findings):
        unknown_evidence = set(item.get("evidence_ids", [])) - evidence_ids
        unknown_substitutes = set(item.get("substitute_ids", [])) - substitute_ids
        unknown_competitors = set(item.get("competitor_slugs", [])) - competitor_slugs
        if unknown_evidence or unknown_substitutes or unknown_competitors:
            reference_errors.append(
                f"evidence={sorted(unknown_evidence)} "
                f"substitutes={sorted(unknown_substitutes)} "
                f"competitors={sorted(unknown_competitors)}"
            )
        if "job_id" in item and item["job_id"] not in SUBSTITUTE_JOB_IDS:
            reference_errors.append(f"unknown Job {item['job_id']}")
        for ref in item.get("market_research_refs", []):
            if ref.startswith("interview_findings."):
                valid = ref.split(".", 1)[1] in market.get("interview_findings", {})
            elif ref.startswith("hypotheses."):
                valid = ref.split(".", 1)[1] in hypothesis_names
            elif ref.startswith("research_inventory."):
                valid = ref.split(".", 1)[1] in inventory_areas
            else:
                valid = False
            if not valid:
                reference_errors.append(f"unknown market-research reference {ref}")
        if "confidence" in item and item["confidence"] not in FINDINGS_CONFIDENCE:
            confidence_errors.append(str(item["confidence"]))
        if material_markers & set(item):
            has_trace = bool(item.get("evidence_ids") or item.get("market_research_refs"))
            explicit_gap = "Insufficient Evidence" in json.dumps(item, ensure_ascii=False)
            if not has_trace and not explicit_gap:
                material_without_trace.append(
                    next(iter(material_markers & set(item)))
                )
    validation.check(
        not reference_errors,
        "active findings resolve every evidence, substitute, Job, competitor, and market reference",
    )
    validation.check(
        not confidence_errors,
        "active findings use qualitative controlled confidence values",
    )
    validation.check(
        not material_without_trace,
        "every material active conclusion is evidence-traceable or explicitly Insufficient Evidence",
    )

    jobs = findings.get("jobs", [])
    validation.check(
        len(jobs) == 4
        and {item.get("job_id") for item in jobs} == SUBSTITUTE_JOB_IDS,
        "active conclusions analyze all four Jobs separately",
    )
    red_team = findings.get("red_team", [])
    validation.check(
        len(red_team) == 10
        and {item.get("id") for item in red_team}
        == {f"RT-{number:02d}" for number in range(1, 11)}
        and all(
            all(
                item.get(key)
                for key in (
                    "supporting_evidence",
                    "counterevidence",
                    "unknowns",
                    "confidence",
                    "customer_discovery_test",
                )
            )
            for item in red_team
        ),
        "all ten required Red Team possibilities retain support, counterevidence, unknowns, confidence, and a test",
    )
    all_keys = set(nested_keys(findings))
    validation.check(
        not any(
            re.search(
                r"(?:score|rating|percentage|threat_rank|priority_score)$",
                key,
                flags=re.IGNORECASE,
            )
            for key in all_keys
        ),
        "active findings define no arbitrary numeric research or decision score",
    )
    raw_text = FINDINGS_PATH.read_text(encoding="utf-8")
    validation.check(
        '"current_decision": "Narrow the research"' in raw_text
        and "does not establish product-market fit" in raw_text
        and "does not establish demand for BizMatch" in raw_text,
        "active findings preserve the skeptical decision and adoption boundaries",
    )

    if not artifacts_required:
        return
    validation.check(
        FINDINGS_PAGE.exists() and FINDINGS_REPORT.exists(),
        "active findings generate both the website section and Markdown report",
    )
    if not FINDINGS_PAGE.exists() or not FINDINGS_REPORT.exists():
        return
    page_text = FINDINGS_PAGE.read_text(encoding="utf-8")
    report_text = FINDINGS_REPORT.read_text(encoding="utf-8")
    required_sections = (
        "Conclusions by Job-to-be-Done",
        "Cross-market findings",
        "Strongest competitive pressures",
        "BizMatch assumptions that gained support",
        "BizMatch assumptions that weakened",
        "Supported pains versus hypotheses",
        "What can and cannot be concluded",
        "Customer Discovery agenda",
        "Conditional strategic implications",
        "Required Red Team assessment",
        "Methodology and evidence limitations",
    )
    validation.check(
        all(section in page_text for section in required_sections)
        and all(section in report_text for section in required_sections),
        "active website and report contain every required conclusions section",
    )
    used_evidence = {
        evidence_id
        for item in walk_dicts(findings)
        for evidence_id in item.get("evidence_ids", [])
    }
    validation.check(
        all(
            f"#evidence-{evidence_id}" in page_text
            and f"[{evidence_id}]" in report_text
            for evidence_id in used_evidence
        ),
        "every cited evidence ID is linked from both active conclusion outputs",
    )
    required_labels = (
        "Evidence suggests",
        "Company Claim",
        "Independent Evidence",
        "Inference",
        "Unverified",
        "Insufficient Evidence",
        "Hypothesis requiring validation",
    )
    validation.check(
        all(label in page_text for label in required_labels),
        "active conclusions visibly distinguish evidence and uncertainty labels",
    )
    validation.check(
        "numeric threat ranks" in page_text
        and "This is a research agenda only" in page_text
        and "product-market fit" in page_text,
        "active conclusions preserve rank, research-agenda, and PMF boundaries",
    )
    full_report_pages = list((ROOT / "sites" / "full-report-site").rglob("*.html"))
    validation.check(
        bool(full_report_pages)
        and all(
            'href="/findings-conclusions/"' in path.read_text(encoding="utf-8")
            and ">Findings</a>" in path.read_text(encoding="utf-8")
            for path in full_report_pages
        )
        and 'aria-current="page">Findings</a>' in page_text,
        "Findings is a keyboard-accessible top-level tab across existing routes",
    )
    netlify_config = (ROOT / "netlify.toml").read_text(encoding="utf-8")
    validation.check(
        'from = "/findings-conclusions"' in netlify_config
        and 'to = "/findings-conclusions/index.html"' in netlify_config,
        "Netlify direct loading and refresh are configured for /findings-conclusions",
    )


def check_xlsx(validation, headers, rows, required):
    if not XLSX_PATH.exists():
        validation.check(not required, "generated XLSX exists")
        return
    workbook = load_workbook(XLSX_PATH, read_only=True, data_only=True)
    validation.check(
        workbook.sheetnames == ["Competitive Tracker"],
        "XLSX contains the canonical sheet only",
    )
    sheet = workbook["Competitive Tracker"]
    xlsx_rows = [[normalized(cell) for cell in row] for row in sheet.iter_rows(values_only=True)]
    csv_rows = [headers] + [[normalized(row.get(header, "")) for header in headers] for row in rows]
    validation.check(len(xlsx_rows) == EXPECTED_ROWS + 1, "XLSX contains 36 data rows")
    validation.check(
        bool(xlsx_rows) and len(xlsx_rows[0]) == EXPECTED_COLUMNS,
        "XLSX contains 65 columns",
    )
    validation.check(xlsx_rows == csv_rows, "CSV and XLSX match after normalization")
    with zipfile.ZipFile(XLSX_PATH) as package:
        core_properties = package.read("docProps/core.xml")
    fixed_timestamp = b"2026-07-30T00:00:00Z"
    validation.check(
        core_properties.count(fixed_timestamp) == 2,
        "XLSX created and modified metadata use fixed timestamps",
    )


def function_source(path, function_name):
    source = path.read_text(encoding="utf-8")
    tree = ast.parse(source)
    for node in tree.body:
        if isinstance(node, ast.FunctionDef) and node.name == function_name:
            return ast.get_source_segment(source, node) or ""
    raise AssertionError(f"Function not found: {function_name}")


def check_scoring(validation, rows, artifacts_required):
    source = function_source(GENERATOR, "relationship_score")
    company_names = [row["company"] for row in rows]
    hardcoded = [name for name in company_names if repr(name) in source or f'"{name}"' in source]
    validation.check(not hardcoded, "scoring contains no hard-coded company-name conditions")
    banned_tokens = (
        "product_overlap_score",
        "feature_maturity_score",
        "market_traction_score",
        "funding_strength_score",
        "ai_depth_score",
        "nda_security_strength_score",
        "network_moat_score",
        "direct_threat_score",
        '"free"',
        "'free'",
    )
    validation.check(
        not any(token in source for token in banned_tokens),
        "relationship score does not consume legacy fields or price keywords",
    )
    validation.check(
        "default" not in source.lower(),
        "relationship score has no numeric missing-data default",
    )
    validation.check(
        '"value": None' in source and '"status": "Insufficient Evidence"' in source,
        "missing score inputs return null / Insufficient Evidence",
    )

    active_code = []
    for base in (ROOT / "tools", ROOT / "sites" / "full-report-site"):
        for path in base.rglob("*"):
            if path.is_file() and path.suffix in {".py", ".js"}:
                active_code.append(path)
    cited_dataset_pattern = re.compile(
        r"bizmatch-competitive-research-cited\.(?:csv|xlsx)"
    )
    cited_refs = []
    for path in active_code:
        for line_number, line in enumerate(path.read_text(encoding="utf-8").splitlines(), 1):
            if cited_dataset_pattern.search(line) and any(
                token in line for token in ("open(", "DictReader", "load_workbook", "read_csv")
            ):
                cited_refs.append(f"{path.relative_to(ROOT)}:{line_number}")
    validation.check(not cited_refs, "active code has no archived cited-dataset reference")

    if not artifacts_required:
        return
    generated_data = (
        ROOT / "sites" / "full-report-site" / "canonical-data.js"
    ).read_text(encoding="utf-8") if (ROOT / "sites" / "full-report-site" / "canonical-data.js").exists() else ""
    ranker_data = (
        ROOT / "sites" / "full-report-site" / "ranker-data.js"
    ).read_text(encoding="utf-8") if (ROOT / "sites" / "full-report-site" / "ranker-data.js").exists() else ""
    validation.check(
        not any(f'"{field}"' in generated_data for field in LEGACY_FIELDS),
        "generated canonical site data excludes all legacy scores",
    )
    validation.check(
        not any(f'"{field}"' in ranker_data for field in LEGACY_FIELDS),
        "active ranker data excludes all legacy scores",
    )
    validation.check(
        not re.search(r'"relationship_score"\s*:\s*(?!null)\d', generated_data),
        "generated relationship scores are null when evidence is insufficient",
    )
    active_pages = [
        ROOT / "sites" / "full-report-site" / name
        for name in (
            "index.html",
            "priority-competitors.html",
            "research-table.html",
            "category-analysis.html",
        )
    ]
    active_pages.extend((ROOT / "sites" / "full-report-site" / "companies").glob("*.html"))
    active_text = "\n".join(
        path.read_text(encoding="utf-8") for path in active_pages if path.exists()
    )
    validation.check(
        not any(f'"{field}"' in active_text or f">{field}<" in active_text for field in LEGACY_FIELDS),
        "active company and analysis pages do not render legacy score fields",
    )
    validation.check(
        not re.search(r"\b(?:High|Medium) threat\b", active_text, flags=re.IGNORECASE),
        "active pages contain no score-derived High/Medium threat labels",
    )


def check_profiles(validation, by_company, required):
    missing = []
    for company, fields in PROFILE_FIELDS.items():
        page = ROOT / "sites" / "full-report-site" / "companies" / f"{slug(company)}.html"
        if not page.exists():
            if required:
                missing.append(f"{company}: page missing")
            continue
        content = page.read_text(encoding="utf-8")
        for field in fields:
            if html.escape(by_company[company][field]) not in content:
                missing.append(f"{company}.{field}")
        if "no numeric relationship or threat score is published" not in content:
            missing.append(f"{company}: scoring boundary")
    validation.check(not missing, "six reconciled company pages reflect canonical values")


def check_profile_manifest(validation, rows, message):
    expected = {f"{slug(row['company'])}.html" for row in rows}
    company_dir = ROOT / "sites" / "full-report-site" / "companies"
    actual = {path.name for path in company_dir.glob("*.html")}
    validation.check(
        actual == expected,
        message,
    )


def check_investor_facing_strategy(validation, required):
    active = ROOT / "sites" / "full-report-site" / "index.html"
    historical = (
        ROOT
        / "sites"
        / "full-report-site"
        / "strategic-conclusions-historical.html"
    )
    if not active.exists() or not historical.exists():
        validation.check(not required, "strategic recommendations are isolated from the active page")
        return
    active_text = active.read_text(encoding="utf-8")
    findings_text = (
        FINDINGS_PAGE.read_text(encoding="utf-8")
        if FINDINGS_PAGE.exists()
        else ""
    )
    historical_text = historical.read_text(encoding="utf-8")
    investor_risk_phrases = (
        "Recommended Israeli Beachhead",
        "Proposed launch sequence",
        "Where BizMatch Can Win",
        "First 100 qualified users",
    )
    validation.check(
        not any(phrase in active_text for phrase in investor_risk_phrases)
        and not any(phrase in findings_text for phrase in investor_risk_phrases)
        and "What BizMatch is really competing against" in active_text
        and "Current research decision" in active_text
        and "Conditional strategic implications" in findings_text
        and "ARCHIVED — not current investor conclusions" in historical_text,
        "strategic recommendations are isolated from the active presentation",
    )


def check_presentation_ux(validation, rows, required):
    active_pages = (
        ROOT / "index.html",
        FINDINGS_PAGE,
        PRESENTATION_PAGE,
        MARKET_RESEARCH_PAGE,
        ROOT / "sites" / "full-report-site" / "category-analysis.html",
        ROOT / "sites" / "full-report-site" / "alternative-workflows.html",
        ROOT / "sites" / "full-report-site" / "research-table.html",
        ROOT / "sites" / "full-report-site" / "sources-methodology.html",
    )
    validation.check(
        all(path.exists() for path in active_pages) if required else True,
        "all presentation-first top-level pages are generated",
    )
    if not required or not all(path.exists() for path in active_pages):
        return

    contents = {path: path.read_text(encoding="utf-8") for path in active_pages}
    findings = json.loads(FINDINGS_PATH.read_text(encoding="utf-8"))
    _, substitute_rows = read_csv_table(SUBSTITUTES_PATH)
    _, evidence_rows = read_csv_table(SUBSTITUTE_EVIDENCE_PATH)
    overview = contents[ROOT / "index.html"]
    presentation = contents[PRESENTATION_PAGE]
    findings_page = contents[FINDINGS_PAGE]
    workflows = contents[ROOT / "sites" / "full-report-site" / "alternative-workflows.html"]
    landscape = contents[ROOT / "sites" / "full-report-site" / "category-analysis.html"]

    validation.check(
        findings["executive_conclusion"]["headline"] in overview
        and findings["executive_conclusion"]["headline"] in presentation
        and str(len(rows)) in overview
        and str(len(substitute_rows)) in overview
        and str(len(evidence_rows)) in overview,
        "overview metrics and central conclusion derive from canonical sources",
    )
    job_titles = [job["title"] for job in findings["jobs"]]
    validation.check(
        all(
            all(title in contents[path] for title in job_titles)
            for path in (ROOT / "index.html", FINDINGS_PAGE, PRESENTATION_PAGE)
        ),
        "all four Jobs appear in the overview, findings, and presentation",
    )
    validation.check(
        all(item["finding"] in presentation for item in findings["cross_market_findings"])
        and all(
            item["assumption"] in presentation
            for item in findings["assumptions_strengthened"]
            + findings["assumptions_weakened"]
        ),
        "presentation claims resolve to active canonical findings",
    )
    validation.check(
        all(
            html.escape(row["name"]) in workflows
            and f'id="substitute-{row["substitute_id"]}"' in workflows
            for row in substitute_rows
        )
        and all(
            html.escape(row["source_title"]) in workflows
            and f'id="evidence-{row["evidence_id"]}"' in workflows
            for row in evidence_rows
        ),
        "workflow explorer uses human labels with resolvable canonical records",
    )
    validation.check(
        all(
            row["company"] in landscape
            and f'companies/{slug(row["company"])}.html' in landscape
            for row in rows
        ),
        "every competitor slug resolves to a human-readable company label",
    )
    raw_primary_label = re.compile(
        r"<(?:a|h[1-6])[^>]*>\s*(?:EV-\d+|SUB-\d+|JOB-[A-Z-]+)\s*</",
        flags=re.IGNORECASE,
    )
    validation.check(
        not any(raw_primary_label.search(content) for content in contents.values()),
        "raw evidence, substitute, and Job IDs are never primary interface labels",
    )
    validation.check(
        'id="substituteSearch"' in workflows
        and 'id="substituteJob"' in workflows
        and 'id="substituteCategory"' in workflows
        and 'id="substituteStrength"' in workflows
        and 'id="substituteStatus"' in workflows
        and 'id="substituteStage"' in workflows
        and all(row["substitute_id"] in workflows for row in substitute_rows),
        "workflow filters derive from the canonical substitute layer",
    )
    validation.check(
        "Company-level competitive relationship" in landscape
        and "Workflow substitute" in landscape
        and "Company-level competitive relationship" in findings_page
        and "Workflow substitutes" in findings_page,
        "company competitors and workflow substitutes are explicitly distinguished",
    )
    validation.check(
        all(
            'class="skip-link"' in content
            and 'id="main-content"' in content
            and 'aria-label="Primary research navigation"' in content
            for content in contents.values()
        ),
        "top-level pages include semantic landmarks and a keyboard skip link",
    )
    css = (ROOT / "sites" / "full-report-site" / "style.css").read_text(encoding="utf-8")
    validation.check(
        ":focus-visible" in css
        and "prefers-reduced-motion" in css
        and "@media print" in css
        and "@media (max-width:700px)" in css,
        "focus, reduced-motion, print, and mobile styles are present",
    )
    validation.check(
        'from = "/presentation"' in (ROOT / "netlify.toml").read_text(encoding="utf-8")
        and 'aria-current="page">Presentation</a>' in presentation
        and "data-presentation-next" in presentation,
        "presentation mode has a stable route and keyboard-friendly controls",
    )

    broken = []
    for page_path, content in contents.items():
        for href in re.findall(r'href="([^"]+)"', content):
            parsed = urlparse(href)
            if parsed.scheme or href.startswith(("mailto:", "data:")):
                continue
            clean = unquote(parsed.path)
            if not clean:
                continue
            target = ROOT / clean.lstrip("/") if clean.startswith("/") else page_path.parent / clean
            if clean.endswith("/"):
                target = target / "index.html"
            if not target.exists():
                broken.append(f"{page_path.relative_to(ROOT)} -> {href}")
    validation.check(not broken, "top-level navigation and contextual links resolve locally")


def artifact_paths():
    paths = [
        ROOT / "index.html",
        ROOT / "START_HERE.html",
        ROOT / "README.html",
        ROOT / "SUBSTITUTE_MATRIX.md",
        ROOT / "SUBSTITUTE_WORKFLOWS.md",
        ROOT / "SUBSTITUTE_EVIDENCE_REGISTER.md",
        FINDINGS_REPORT,
        MARKET_RESEARCH_PAGE,
        FINDINGS_PAGE,
        XLSX_PATH,
    ]
    for base in (
        ROOT / "sites" / "full-report-site",
        ROOT / "sites" / "citation-site",
        ROOT / "reports",
        ROOT / "presentation",
    ):
        paths.extend(
            path
            for path in base.rglob("*")
            if path.is_file() and path.suffix in {".html", ".js", ".md"}
        )
    return sorted(set(paths))


def hashes():
    return {
        str(path.relative_to(ROOT)): hashlib.sha256(path.read_bytes()).hexdigest()
        for path in artifact_paths()
        if path.exists()
    }


def run_build_twice(validation):
    commands = (
        [sys.executable, str(ROOT / "tools" / "build_xlsx.py")],
        [sys.executable, str(SUBSTITUTE_REPORT_SCRIPT)],
        [sys.executable, str(FINDINGS_REPORT_SCRIPT)],
        [sys.executable, str(GENERATOR)],
    )
    before = hashes()
    for command in commands:
        subprocess.run(command, cwd=ROOT, check=True)
    first = hashes()
    time.sleep(2.1)
    for command in commands:
        subprocess.run(command, cwd=ROOT, check=True)
    second = hashes()
    validation.check(
        first == second,
        "time-separated builds are byte-for-byte deterministic",
    )
    validation.check(
        before == second,
        "checked-in generated artifacts match the deterministic build",
    )


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--pre-build", action="store_true", help="skip required derived-artifact checks")
    parser.add_argument("--check-build", action="store_true", help="run the build twice and compare hashes")
    args = parser.parse_args()

    validation = Validation()
    headers, rows = read_canonical()
    by_company = check_canonical(validation, headers, rows)
    check_substitutes(validation, rows, artifacts_required=not args.pre_build)
    check_market_research(validation, artifacts_required=not args.pre_build)
    check_findings(validation, rows, artifacts_required=not args.pre_build)
    check_scoring(validation, rows, artifacts_required=not args.pre_build)
    if not args.pre_build:
        check_profile_manifest(
            validation,
            rows,
            "company profile directory contains exactly the canonical profiles",
        )
        check_investor_facing_strategy(validation, required=True)
        check_presentation_ux(validation, rows, required=True)
    if args.check_build:
        run_build_twice(validation)
        check_profile_manifest(
            validation,
            rows,
            "site build emits no extra company-profile files",
        )
    if not args.pre_build:
        check_xlsx(validation, headers, rows, required=True)
        check_profiles(validation, by_company, required=True)
    validation.require()
    print(f"PASS: {len(validation.passes)} integrity checks")
    for message in validation.passes:
        print(f"  - {message}")


if __name__ == "__main__":
    main()
